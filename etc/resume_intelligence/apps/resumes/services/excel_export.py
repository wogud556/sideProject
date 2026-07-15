"""이력서 분석 결과를 openpyxl 워크북(.xlsx)으로 내보낸다."""
from io import BytesIO

from openpyxl import Workbook

from apps.pdf_analysis.schemas.profile import ResumeProfile

from .. import permissions, repository


def _write_rows(ws, header: list[str], rows: list[list]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(row)


def build_resume_workbook(profile: ResumeProfile) -> Workbook:
    wb = Workbook()

    ws_basic = wb.active
    ws_basic.title = "기본정보"
    basic = profile.basic
    _write_rows(
        ws_basic,
        ["항목", "값", "신뢰도", "검토필요"],
        [
            ["이름", basic.name.value, basic.name.confidence, basic.name.review_required],
            ["영문이름", basic.name_en.value, basic.name_en.confidence, basic.name_en.review_required],
            ["생년월일", basic.birth_date.value, basic.birth_date.confidence, basic.birth_date.review_required],
            ["이메일", basic.email.value, basic.email.confidence, basic.email.review_required],
            ["전화", basic.phone.value, basic.phone.confidence, basic.phone.review_required],
            ["주소", basic.address.value, basic.address.confidence, basic.address.review_required],
            ["포트폴리오", basic.portfolio_url.value, basic.portfolio_url.confidence, basic.portfolio_url.review_required],
            ["GitHub", basic.github_url.value, basic.github_url.confidence, basic.github_url.review_required],
            ["LinkedIn", basic.linkedin_url.value, basic.linkedin_url.confidence, basic.linkedin_url.review_required],
            ["요약", basic.summary.value, basic.summary.confidence, basic.summary.review_required],
        ],
    )

    ws_careers = wb.create_sheet("경력")
    _write_rows(
        ws_careers,
        ["회사명", "부서", "직급", "시작일", "종료일", "재직중", "주요업무", "신뢰도", "검토필요"],
        [
            [
                c.company_name_raw,
                c.department,
                c.position,
                c.start_date,
                c.end_date,
                c.is_current,
                "; ".join(c.responsibilities),
                c.confidence,
                c.review_required,
            ]
            for c in profile.careers
        ],
    )

    ws_edu = wb.create_sheet("학력")
    _write_rows(
        ws_edu,
        ["학교명", "전공", "학위", "시작일", "종료일", "졸업상태", "신뢰도", "검토필요"],
        [
            [e.school_name, e.major, e.degree, e.start_date, e.end_date, e.graduation_status, e.confidence, e.review_required]
            for e in profile.educations
        ],
    )

    ws_proj = wb.create_sheet("프로젝트")
    _write_rows(
        ws_proj,
        ["프로젝트명", "소속", "시작일", "종료일", "역할", "기술스택", "신뢰도", "검토필요"],
        [
            [
                p.project_name,
                p.organization,
                p.start_date,
                p.end_date,
                p.role,
                ", ".join(p.technologies),
                p.confidence,
                p.review_required,
            ]
            for p in profile.projects
        ],
    )

    ws_cert = wb.create_sheet("자격증")
    _write_rows(
        ws_cert,
        ["자격증명", "발급기관", "취득일", "신뢰도", "검토필요"],
        [
            [c.certificate_name, c.issuer, c.acquired_date, c.confidence, c.review_required]
            for c in profile.certificates
        ],
    )

    ws_skill = wb.create_sheet("기술")
    _write_rows(
        ws_skill,
        ["분류", "이름", "수준", "신뢰도", "검토필요"],
        [[s.category, s.name, s.level, s.confidence, s.review_required] for s in profile.skills],
    )

    ws_lang = wb.create_sheet("어학")
    _write_rows(
        ws_lang,
        ["언어", "시험명", "점수", "등급", "취득일", "신뢰도", "검토필요"],
        [
            [lg.language, lg.test_name, lg.score, lg.grade, lg.acquired_date, lg.confidence, lg.review_required]
            for lg in profile.languages
        ],
    )

    return wb


def export_resume_to_excel(resume_id, user_id: str) -> BytesIO:
    permissions.get_owned_document_or_404(resume_id, user_id)
    profile_data = repository.read_profile(resume_id) or {}
    profile = ResumeProfile.model_validate(profile_data)

    workbook = build_resume_workbook(profile)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
