import shutil
import tempfile
from pathlib import Path

from django.core.files.uploadedfile import SimpleUploadedFile
from django.http import Http404
from django.test import Client, SimpleTestCase, override_settings
from openpyxl import load_workbook

from apps.pdf_analysis.schemas.profile import ResumeProfile
from apps.resumes import repository
from apps.resumes.services import excel_export, upload_service

from .helpers import build_text_resume_pdf, signup_and_login


class BuildResumeWorkbookTests(SimpleTestCase):
    def test_creates_expected_sheets(self):
        profile = ResumeProfile()
        wb = excel_export.build_resume_workbook(profile)
        self.assertEqual(wb.sheetnames, ["기본정보", "경력", "학력", "프로젝트", "자격증", "기술", "어학"])

    def test_career_rows_contain_data(self):
        profile = ResumeProfile.model_validate(
            {
                "careers": [
                    {
                        "company_name_raw": "테스트회사",
                        "start_date": "2020-01",
                        "end_date": "2021-01",
                        "responsibilities": ["업무1", "업무2"],
                        "confidence": 0.9,
                    }
                ]
            }
        )
        wb = excel_export.build_resume_workbook(profile)
        ws = wb["경력"]
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "회사명")
        self.assertEqual(rows[1][0], "테스트회사")
        self.assertEqual(rows[1][6], "업무1; 업무2")


class ExportResumeToExcelTests(SimpleTestCase):
    def setUp(self):
        self.data_dir = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, self.data_dir, ignore_errors=True)
        self.override = override_settings(DATA_DIR=self.data_dir)
        self.override.enable()
        self.addCleanup(self.override.disable)

        upload = SimpleUploadedFile("resume.pdf", build_text_resume_pdf(), content_type="application/pdf")
        self.resume_id = upload_service.upload_resume("user-1", upload)

    def test_returns_loadable_xlsx_bytes(self):
        buffer = excel_export.export_resume_to_excel(self.resume_id, "user-1")
        wb = load_workbook(buffer)
        ws = wb["기본정보"]
        rows = list(ws.iter_rows(values_only=True))
        name_row = next(r for r in rows if r[0] == "이름")
        self.assertEqual(name_row[1], "홍길동")

    def test_non_owner_raises_404(self):
        with self.assertRaises(Http404):
            excel_export.export_resume_to_excel(self.resume_id, "someone-else")

    def test_all_sheets_match_profile_json_content(self):
        """SAMPLE_RESUME_LINES(경력/학력/프로젝트/자격증/기술/어학 모두 포함)로
        생성한 profile.json의 각 섹션 값이 Excel 시트 데이터와 정확히 일치하는지
        확인한다(시트별 데이터 정합성)."""
        profile_data = repository.read_profile(self.resume_id)
        self.assertTrue(profile_data["careers"], "샘플 이력서에 경력이 추출되지 않았습니다")
        self.assertTrue(profile_data["educations"])
        self.assertTrue(profile_data["projects"])
        self.assertTrue(profile_data["certificates"])
        self.assertTrue(profile_data["skills"])
        self.assertTrue(profile_data["languages"])

        buffer = excel_export.export_resume_to_excel(self.resume_id, "user-1")
        wb = load_workbook(buffer)

        basic_rows = {row[0]: row[1] for row in wb["기본정보"].iter_rows(values_only=True)}
        self.assertEqual(basic_rows["이름"], profile_data["basic"]["name"]["value"])
        self.assertEqual(basic_rows["이메일"], profile_data["basic"]["email"]["value"])
        self.assertEqual(basic_rows["전화"], profile_data["basic"]["phone"]["value"])

        career_rows = list(wb["경력"].iter_rows(values_only=True))[1:]
        self.assertEqual(len(career_rows), len(profile_data["careers"]))
        for row, career in zip(career_rows, profile_data["careers"]):
            self.assertEqual(row[0], career["company_name_raw"])
            self.assertEqual(row[3], career["start_date"])
            self.assertEqual(row[4], career["end_date"])
            self.assertEqual(row[5], career["is_current"])
            self.assertEqual(row[6], "; ".join(career["responsibilities"]))

        edu_rows = list(wb["학력"].iter_rows(values_only=True))[1:]
        self.assertEqual(len(edu_rows), len(profile_data["educations"]))
        for row, edu in zip(edu_rows, profile_data["educations"]):
            self.assertEqual(row[0], edu["school_name"])
            self.assertEqual(row[1], edu["major"])
            self.assertEqual(row[3], edu["start_date"])
            self.assertEqual(row[4], edu["end_date"])

        proj_rows = list(wb["프로젝트"].iter_rows(values_only=True))[1:]
        self.assertEqual(len(proj_rows), len(profile_data["projects"]))
        for row, proj in zip(proj_rows, profile_data["projects"]):
            self.assertEqual(row[0], proj["project_name"])
            self.assertEqual(row[2], proj["start_date"])
            self.assertEqual(row[3], proj["end_date"])

        cert_rows = list(wb["자격증"].iter_rows(values_only=True))[1:]
        self.assertEqual(len(cert_rows), len(profile_data["certificates"]))
        for row, cert in zip(cert_rows, profile_data["certificates"]):
            self.assertEqual(row[0], cert["certificate_name"])
            self.assertEqual(row[1], cert["issuer"])
            self.assertEqual(row[2], cert["acquired_date"])

        skill_rows = list(wb["기술"].iter_rows(values_only=True))[1:]
        self.assertEqual(len(skill_rows), len(profile_data["skills"]))
        for row, skill in zip(skill_rows, profile_data["skills"]):
            self.assertEqual(row[1], skill["name"])

        lang_rows = list(wb["어학"].iter_rows(values_only=True))[1:]
        self.assertEqual(len(lang_rows), len(profile_data["languages"]))
        for row, lang in zip(lang_rows, profile_data["languages"]):
            self.assertEqual(row[0], lang["language"])
            self.assertEqual(row[1], lang["test_name"])
            self.assertEqual(row[2], lang["score"])
            self.assertEqual(row[4], lang["acquired_date"])


class ExcelDownloadViewTests(SimpleTestCase):
    def setUp(self):
        self.data_dir = Path(tempfile.mkdtemp())
        self.addCleanup(shutil.rmtree, self.data_dir, ignore_errors=True)
        self.override = override_settings(DATA_DIR=self.data_dir)
        self.override.enable()
        self.addCleanup(self.override.disable)

        self.client = Client()
        self.user = signup_and_login(self.client)
        upload = SimpleUploadedFile("resume.pdf", build_text_resume_pdf(), content_type="application/pdf")
        self.resume_id = upload_service.upload_resume(self.user["id"], upload)

    def test_template_view_downloads_xlsx(self):
        response = self.client.get(f"/resumes/{self.resume_id}/excel/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])

    def test_api_view_downloads_xlsx(self):
        response = self.client.get(f"/api/v1/resumes/{self.resume_id}/excel")
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])

    def test_non_owner_gets_404(self):
        other_client = Client()
        signup_and_login(other_client, username="other3")
        response = other_client.get(f"/resumes/{self.resume_id}/excel/")
        self.assertEqual(response.status_code, 404)
